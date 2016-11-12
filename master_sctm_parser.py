import jinja2
import openpyxl
import os
import string
import sys

from ntmap import *

# Get the path of this file
dir_path = os.path.dirname(os.path.realpath(__file__))

# Load the Jinja2 template
loader = jinja2.FileSystemLoader(dir_path)
environment = jinja2.Environment(loader=loader)
controls_template = environment.get_template('security_control.j2')
crm_template = environment.get_template('crm.j2')

# Open the SCTM Excel workbook rw so we can look at row dimensions.
wb = openpyxl.load_workbook(sys.argv[1], read_only=False)
ws = wb['Matrix']

# Build the column index
column_index = {}
for row in ws.iter_rows(min_row=1, max_row=1):
    for i, cell in enumerate(row):
        column_index[i] = cell.value

# Init the sctm dict, to be keyed to control title
sctm = {}

# Iterate over every row in the sheet, starting at the 2nd row
for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    # Skip hidden rows
    if ws.row_dimensions[i].hidden is True:
        continue

    # Temporary dict for this iteration
    r = {}

    # Iterate over the cells and insert the values into the temporary dict r
    # keyed with the column name.
    for j, cell in enumerate(row):
        # If the cell is non-blank, insert the cell's contents into the temp
        # dict r keyed by column name.
        if cell.value and cell.value != '':
            r[column_index[j]] = cell.value
        else:
            r[column_index[j]] = 'undefined'

    # Key for the temp dict
    ref_num = r['NIST Ref #']

    # Build the control title with nice capitalization
    r["Title"] = " ".join(w.capitalize() for w in str(nt_map[ref_num]).split())

    # First we check to see if this control has been processed before
    # Remove all those pesky new lines
    r['Requirements'] = r['Requirements'].replace(u"\n", u" ")
    if sctm.get(ref_num):
        # If it has, we concat the control language
        sctm[ref_num][0]['Original Requirements'] += (u" " + r['Requirements'])
    else:
        # Otherwise we start fresh
        r['Original Requirements'] = r['Requirements']

    # Build an alphabetic part number to jive with the FedRAMP template
    r['Part'] = string.lowercase[int(r['Org Ref #'].split("_")[-1])]

    # Now we build a list of control parts and store it under the same key
    if sctm.get(ref_num):
        sctm[ref_num].append(r)
    else:
        sctm[ref_num] = [r]

# Process the master sctm data structure and look for Tenant responsible
# controls
crm = {}
for key in sctm.keys():
    for part in sctm[key]:
        if u"OpenShift Tenant" in part['Role']:
            crm[key] = sctm[key]

output = controls_template.render(sctm=sctm)
with open('controls.rst', 'w') as controls_file:
    controls_file.write(output.encode('utf-8'))

output = crm_template.render(sctm=crm)
with open('crm.rst', 'w') as controls_file:
    controls_file.write(output.encode('utf-8'))

